from openpyxl import Workbook, load_workbook
import pandas as pd



planilha = load_workbook("Text.xlsx", data_only = True)

aba_ativa = planilha.active


ws = planilha['my test']


sheet_obj = planilha.active
row = sheet_obj.max_row
column = sheet_obj.max_column


album = {'nome':'A Night at""" the Opera,' ' TESTE1,' ' TESTE2',' artista':'Blind Guardian','lançamento':2002}




#LOCALIZANDO QUANTOS DADOS TEM NA LINHA E EM QUAL LINHA FOI ENCONTRADO
i=0
index= 0
b=0
list= []

for celula in aba_ativa["F"]:
        i+= 1
        if celula.value == "Gabriel":
            sheet_obj.cell(row = i, column = 7)
            list.append(i)
            index +=1

print(f'Quantidade de Valores Encontrados na tabela: {index}')
print(f'O valor foi encontrado na linha: {list}')


#Consultando a tabela toda
# df = pd.read_excel("Text.xlsx")
# print(df)



# SELECIONANDO UMA PLANILHA COM PANDAS
#                               *sheet_name* Escolhe planilha dentro do arq
# df2 = pd.read_excel("Text.xlsx", sheet_name=1, usecols=None)
# print(df2)



#SELECIONANDO UMA COLUNA COM PANDAS
# df2 = pd.read_excel("Text.xlsx", sheet_name=1, usecols="A")
# print(df2)



#SELECIONADO UMA COLUNA PULANDO linhas  COM PANDAS
# df3 = pd.read_excel("Text.xlsx", sheet_name=1,skiprows=2, usecols="A")
# print(df3)



#SELECIONADO UMA LINHA COM PANDAS
# df4 = pd.read_excel("Text.xlsx", sheet_name=1,nrows=1)
# print(df4)



#ws['G1'] = '=SUM(F1:F13)'
#ws['G3'] = '=AVERAGE(F1:F13)'



# for i in range(1, row + 1):
#       cell_obj = sheet_obj.cell(row = i, column = 7)
#       print(cell_obj.value)



# for celula in aba_ativa["G"]:
#       if celula.value == "oi":
#          linha = celula.row
#          aba_ativa[f"G{linha}"] = ""


planilha.save("Text.xlsx")



# teste = ws['G1'].value
#print(teste)



#print(teste)
#print(diciTeste)


#calculando a Media
# df = pd.DataFrame(diciTeste)
# mean_df = df['dados'].mean()
# print(mean_df)


#PERCORRENDO td um dicionario
# for key in album:
#    print(album[key])


#print(album['nome'])
list = []

#list.append(album['nome'])


#print(list)


#Inserção de dados simples
#ws['E1'] = "inserção"



#Consulta de dados simples
teste = ws['E1'].value
#print(teste)




# for i in range(1, row + 1):
#      cell_obj = sheet_obj.cell(row = i, column = 5)
#      print(cell_obj.value)



# INSERINDO DADOS COMPARANDO

# for celula in aba_ativa["I"]:
#      if celula.value == "oi":
#         linha = celula.row
#         aba_ativa[f"E{linha}"] = "uu"

print(f'\n')

#CONSULTADO DADOS  Quant de Dados
# index= 0
# for celula in aba_ativa["F"]:
#      if celula.value == 600:
#             index +=1
# print(index)



#album = {'nome':'A Night at the Opera','artista':'Blind Guardian','lançamento':2002}


#função range()

# teste = range(1, 10, +2)
# print(f'\n')
# for x in teste:
#     print(x)



# sheet_obj = planilha.active


# row = sheet_obj.max_row
# column = sheet_obj.max_column


# print("Total Rows:", row)
# print("Total Columns:", column)


#  SELECIONADO  COLUNA
# for i in range(1, row + 1):
#     cell_obj = sheet_obj.cell(row = i, column = 1)
#     print(cell_obj.value)


# # SELECIONANDO  LINHA
# for i in range(1, column + 1):
#     cell_obj = sheet_obj.cell(row = 1, column = i)
#     print(cell_obj.value, end = " | ")
